import openpyxl


def format_spreadsheet(file_path, new_file_name):
    wb = openpyxl.load_workbook(file_path)

    ws = wb.active

    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = ws.title
    model_line = ()
    for row in range(1, ws.max_row + 1):
        current_line = [ws.cell(row, i).value for i in range(1, ws.max_column + 1)]
        print(current_line)
        if ws.cell(row, 1).value:
            model_line = [ws.cell(row, i).value for i in range(1, 4)]
            print(model_line)
            ws1.append(current_line)
        else:
            line_to_write = tuple(model_line) + tuple(current_line[3:])
            ws1.append(line_to_write)
    wb1.save(new_file_name)
